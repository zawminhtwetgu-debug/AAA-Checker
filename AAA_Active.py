import sys
import os

# Ensure Playwright finds the browser binaries when bundled as .exe
if getattr(sys, 'frozen', False):
    os.environ['PLAYWRIGHT_BROWSERS_PATH'] = os.path.join(sys._MEIPASS, 'ms-playwright')

# Load variables from a local .env file if present (for local/dev runs).
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from concurrent.futures import ThreadPoolExecutor
import threading
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import tkinter.font as tkFont

LOGIN_URL = os.environ.get("AAA_LOGIN_URL", "http://10.201.1.160/metro/aaacheck.asp")
USERNAME = os.environ.get("AAA_USERNAME")
PASSWORD = os.environ.get("AAA_PASSWORD")

if not USERNAME or not PASSWORD:
    import tkinter.messagebox as _mb
    _root_check = tk.Tk()
    _root_check.withdraw()
    _mb.showerror(
        "Missing credentials",
        "AAA_USERNAME and AAA_PASSWORD environment variables are not set.\n"
        "Please set them before running this app (see README)."
    )
    sys.exit(1)

results = []
stop_flag = False
total_to_check = 0
checked_count = 0
lock = threading.Lock()
last_sorted_column = None
sort_ascending = True

def scrape_user(username):
    data = {
        'Username': username,
        'Login Status': 'Unknown',
        'Login Fail': 'Unknown error',
        'Login Date': '',
        'Port RX Power': '',
        'Port ID': '',
        'Site ID': '',
        'RX Power Status': '',
        'Status': '',
        'Active Date': '',
        'Suspend Date': ''
    }

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context()
            context.route("**/*", lambda route, request: route.abort() if request.resource_type in ["image", "stylesheet", "font"] else route.continue_())
            page = context.new_page()
            page.goto(LOGIN_URL, timeout=15000)

            page.fill('input[name="username"]', USERNAME)
            page.fill('input[name="password"]', PASSWORD)
            page.click('input[type="submit"]')
            page.wait_for_selector('input[name="username"]', timeout=5000)

            page.fill('input[name="username"]', username)
            page.click('input[name="Check"]')

            try:
                page.wait_for_selector('img[src="images/online_1.png"], img[src="images/offline_1.png"]', timeout=5000)
                if page.is_visible('img[src="images/online_1.png"]'):
                    data['Login Status'] = 'Online'
                    data['Login Fail'] = ''
                elif page.is_visible('img[src="images/offline_1.png"]'):
                    data['Login Status'] = 'Offline'
                    data['Login Fail'] = ''
            except PlaywrightTimeoutError:
                pass

            if page.is_visible('span.style15'):
                data['Login Fail'] = page.inner_text('span.style15').strip()

            # Scrape the three new fields using provided XPaths
            try:
                data['Status'] = page.locator('xpath=/html/body/form/table/tbody/tr[13]/td[2]').inner_text().strip()
                data['Active Date'] = page.locator('xpath=/html/body/form/table/tbody/tr[14]/td[2]/strong').inner_text().strip()
                data['Suspend Date'] = page.locator('xpath=/html/body/form/table/tbody/tr[16]/td[2]').inner_text().strip()
            except:
                pass

            strong_tags = page.query_selector_all('strong')
            for tag in strong_tags:
                text = tag.inner_text().strip()
                if "OLT" in text and ":" in text:
                    data['Port ID'] = text
                    data['Site ID'] = text[:7]
                elif "/" in text and ":" in text:
                    data['Login Date'] = text

            if page.is_visible('td:has-text("OLT:")'):
                data['Port RX Power'] = page.inner_text('td:has-text("OLT:")')

            # Detect RX Power Status
            html = page.content()
            if "Rx power is not good" in html:
                data['RX Power Status'] = "Rx power is not good"
            elif "Rx power is good" in html:
                data['RX Power Status'] = "Rx power is good"

            browser.close()
            return data

    except Exception as e:
        print(f"Error for {username}: {e}")
        return data

def start_scraping():
    global stop_flag, total_to_check, checked_count
    stop_flag = False
    clear_results(clear_input=False)
    usernames = user_input.get("1.0", tk.END).strip().splitlines()
    if not usernames:
        messagebox.showwarning("Warning", "Please enter usernames.")
        return

    total_to_check = len(usernames)
    checked_count = 0
    update_progress()

    def run_scraping():
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = [executor.submit(scrape_and_store, username) for username in usernames]
            for future in futures:
                future.result()
        root.after(100, lambda: messagebox.showinfo("Completed", "All usernames checked."))

    threading.Thread(target=run_scraping).start()

def scrape_and_store(username):
    global checked_count
    if stop_flag:
        return
    data = scrape_user(username)
    with lock:
        existing = next((r for r in results if r['Username'] == username), None)
        if existing:
            results.remove(existing)
        results.append(data)
        checked_count += 1
    update_treeview()
    update_progress()

def stop_scraping():
    global stop_flag
    stop_flag = True

def clear_results(clear_input=True):
    tree.delete(*tree.get_children())
    results.clear()
    progress_bar['value'] = 0
    progress_label.config(text="Progress: 0/0 (0%)")
    if clear_input:
        user_input.delete("1.0", tk.END)

def update_progress():
    percent = int((checked_count / total_to_check) * 100) if total_to_check else 0
    progress_bar['value'] = percent
    progress_label.config(text=f"Progress: {checked_count}/{total_to_check} ({percent}%)")
    root.after(100, update_progress)

def export_to_excel():
    if not results:
        messagebox.showinfo("No data", "No data to export.")
        return
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], initialfile="AAA Checker")
    if file_path:
        df = pd.DataFrame(results)
        df.to_excel(file_path, index=False)

        wb = load_workbook(file_path)
        ws = wb.active

        border = Border(left=Side(border_style="thin"),
                        right=Side(border_style="thin"),
                        top=Side(border_style="thin"),
                        bottom=Side(border_style="thin"))

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(file_path)
        messagebox.showinfo("Success", f"Data exported to {file_path}")

def filter_table(*args):
    query = search_var.get().lower()
    tree.delete(*tree.get_children())
    for row in results:
        if any(query in str(value).lower() for value in row.values()):
            tree.insert("", tk.END, values=[row[col] for col in columns])

def sort_column(col, reverse=False):
    global last_sorted_column, sort_ascending
    if col == last_sorted_column:
        sort_ascending = not sort_ascending
    else:
        sort_ascending = True
    last_sorted_column = col
    sorted_results = sorted(results, key=lambda x: str(x[col]), reverse=not sort_ascending)
    update_treeview(sorted_results)

def update_treeview(sorted_results=None):
    if not sorted_results:
        sorted_results = results
    tree.delete(*tree.get_children())
    for row in sorted_results:
        tree.insert("", tk.END, values=[row[col] for col in columns])
    tree.yview_moveto(1)

# ================= GUI =================

root = tk.Tk()
root.title("AAA Checker")
root.geometry("1200x700")

style = ttk.Style()
style.configure("Treeview.Heading", font=(None, 10, 'bold'))

tk.Label(root, text="AAA Checker", bg="#4287f5", fg="white", font=("Arial", 16, "bold"), padx=10, pady=10).pack(fill=tk.X)
tk.Label(root, text="Created by Zaw Min Htwe", font=("Arial", 10)).pack()

frame = tk.Frame(root)
frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

user_input = tk.Text(frame, height=8, width=30)
user_input.pack(pady=10)
user_input.configure(font=("Arial", 10))

btn_frame = tk.Frame(frame)
btn_frame.pack(pady=5)
tk.Button(btn_frame, text="Start Check", command=start_scraping, bg="#4CAF50", fg="white", width=15).pack(side=tk.LEFT, padx=5)
tk.Button(btn_frame, text="Stop Check", command=stop_scraping, bg="#f44336", fg="white", width=15).pack(side=tk.LEFT, padx=5)
tk.Button(btn_frame, text="Clear Result", command=lambda: clear_results(clear_input=True), bg="#607D8B", fg="white", width=15).pack(side=tk.LEFT, padx=5)
tk.Button(btn_frame, text="Export to Excel", command=export_to_excel, bg="#FF9800", fg="white", width=15).pack(side=tk.LEFT, padx=5)

search_frame = tk.Frame(frame)
search_frame.pack(pady=10)
search_var = tk.StringVar()
search_var.trace("w", filter_table)
tk.Label(search_frame, text="Search Filter", font=("Arial", 10)).pack()
tk.Entry(search_frame, textvariable=search_var, width=30).pack()

# Added the new columns to the display list
columns = ["Username", "Login Status", "Login Fail", "Status", "Active Date", "Suspend Date", "Login Date", "Port RX Power", "Port ID", "Site ID", "RX Power Status"]
tree = ttk.Treeview(frame, columns=columns, show="headings")
for col in columns:
    tree.heading(col, text=col, command=lambda c=col: sort_column(c))
    tree.column(col, anchor="center", width=110)
tree.pack(fill=tk.BOTH, expand=True)

scrollbar_y = ttk.Scrollbar(tree, orient="vertical", command=tree.yview)
tree.configure(yscroll=scrollbar_y.set)
scrollbar_y.pack(side="right", fill="y")

progress_bar = ttk.Progressbar(root, orient="horizontal", length=500, mode="determinate")
progress_bar.pack(pady=10)
progress_label = tk.Label(root, text="Progress: 0/0 (0%)")
progress_label.pack()

root.mainloop()

"""
AAA Checker Active - Streamlit Web App
Converted from Tkinter desktop app by Zaw Min Htwe
Run with: streamlit run streamlit_app.py
"""
import streamlit as st
import pandas as pd
import threading
import time
import io
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

# ─── Configuration ──────────────────────────────────────────────────────
LOGIN_URL = "http://10.201.1.160/metro/aaacheck.asp"
USERNAME = "VMY011432"
PASSWORD = "qaz123"

columns = [
    "Username", "Login Status", "Login Fail", "Status",
    "Active Date", "Suspend Date", "Login Date", "Port RX Power",
    "Port ID", "Site ID", "RX Power Status"
]

# ─── Page config ────────────────────────────────────────────────────────
st.set_page_config(page_title="AAA Checker Active", layout="wide")
st.markdown("""
    <style>
        .stApp { background-color: #f5f5f5; }
        .main-header {
            background: linear-gradient(90deg, #4287f5, #2b6edb);
            color: white; padding: 1.2rem 2rem; border-radius: 10px;
            margin-bottom: 1.5rem;
        }
        .main-header h1 { margin: 0; font-size: 1.8rem; }
        .main-header p { margin: 0; opacity: 0.9; font-size: 0.9rem; }
        div[data-testid="stButton"] button[kind="primary"] {
            background: #4CAF50; color: white; font-weight: bold;
        }
        .stProgress > div > div > div > div { background: #4287f5; }
    </style>
""", unsafe_allow_html=True)

# ─── Initialize session state ──────────────────────────────────────────
if "results" not in st.session_state:
    st.session_state.results = []
if "checked_count" not in st.session_state:
    st.session_state.checked_count = 0
if "total_to_check" not in st.session_state:
    st.session_state.total_to_check = 0
if "scraping_active" not in st.session_state:
    st.session_state.scraping_active = False
if "stop_flag" not in st.session_state:
    st.session_state.stop_flag = False
if "status_message" not in st.session_state:
    st.session_state.status_message = "Ready"

# ─── Scraping functions (same as original) ──────────────────────────────
def scrape_user(username):
    """Scrape data for a single username using Playwright."""
    data = {col: "" for col in columns}
    data["Username"] = username
    data["Login Fail"] = "Unknown error"
    data["Login Status"] = "Unknown"

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context()
            context.route("**/*", lambda route, request: route.abort()
                          if request.resource_type in ["image", "stylesheet", "font"]
                          else route.continue_())
            page = context.new_page()
            page.goto(LOGIN_URL, timeout=15000)
            page.fill('input[name="username"]', USERNAME)
            page.fill('input[name="password"]', PASSWORD)
            page.click('input[type="submit"]')
            page.wait_for_selector('input[name="username"]', timeout=5000)
            page.fill('input[name="username"]', username)
            page.click('input[name="Check"]')

            try:
                page.wait_for_selector(
                    'img[src="images/online_1.png"], img[src="images/offline_1.png"]',
                    timeout=5000
                )
                if page.is_visible('img[src="images/online_1.png"]'):
                    data["Login Status"] = "Online"
                    data["Login Fail"] = ""
                elif page.is_visible('img[src="images/offline_1.png"]'):
                    data["Login Status"] = "Offline"
                    data["Login Fail"] = ""
            except PlaywrightTimeoutError:
                pass

            if page.is_visible("span.style15"):
                data["Login Fail"] = page.inner_text("span.style15").strip()

            try:
                data["Status"] = page.locator(
                    "xpath=/html/body/form/table/tbody/tr[13]/td[2]"
                ).inner_text().strip()
                data["Active Date"] = page.locator(
                    "xpath=/html/body/form/table/tbody/tr[14]/td[2]/strong"
                ).inner_text().strip()
                data["Suspend Date"] = page.locator(
                    "xpath=/html/body/form/table/tbody/tr[16]/td[2]"
                ).inner_text().strip()
            except Exception:
                pass

            for tag in page.query_selector_all("strong"):
                text = tag.inner_text().strip()
                if "OLT" in text and ":" in text:
                    data["Port ID"] = text
                    data["Site ID"] = text[:7]
                elif "/" in text and ":" in text:
                    data["Login Date"] = text

            if page.is_visible("td:has-text('OLT:')"):
                data["Port RX Power"] = page.inner_text("td:has-text('OLT:')")

            html = page.content()
            if "Rx power is not good" in html:
                data["RX Power Status"] = "Rx power is not good"
            elif "Rx power is good" in html:
                data["RX Power Status"] = "Rx power is good"

            browser.close()
            return data
    except Exception as e:
        print(f"Error for {username}: {e}")
        return data


def scrape_and_store(username, results_list, lock):
    """Scrape one user and store result in shared list."""
    if st.session_state.stop_flag:
        return
    data = scrape_user(username)
    with lock:
        # Remove existing entry for same username
        for i, r in enumerate(results_list):
            if r["Username"] == username:
                results_list.pop(i)
                break
        results_list.append(data)
        st.session_state.checked_count += 1


def run_batch(usernames):
    """Run scraping on all usernames using thread pool."""
    lock = threading.Lock()
    st.session_state.results = []
    st.session_state.checked_count = 0
    st.session_state.total_to_check = len(usernames)
    st.session_state.stop_flag = False
    st.session_state.scraping_active = True
    st.session_state.status_message = f"Checking {len(usernames)} users..."

    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [
            executor.submit(scrape_and_store, username, st.session_state.results, lock)
            for username in usernames
        ]
        for f in futures:
            f.result()
            if st.session_state.stop_flag:
                break

    st.session_state.scraping_active = False
    if st.session_state.stop_flag:
        st.session_state.status_message = "Stopped by user"
    else:
        st.session_state.status_message = "Completed ✓"


def generate_excel(results):
    """Generate Excel file in memory and return bytes."""
    df = pd.DataFrame(results)
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False)
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb.active

    border = Border(
        left=Side(border_style="thin"), right=Side(border_style="thin"),
        top=Side(border_style="thin"), bottom=Side(border_style="thin"),
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ─── UI ─────────────────────────────────────────────────────────────────
st.markdown('<div class="main-header"><h1>🔍 AAA Checker</h1>'
            '<p>Created by Zaw Min Htwe</p></div>', unsafe_allow_html=True)

# Status bar
status_col1, status_col2 = st.columns([3, 1])
with status_col1:
    status_placeholder = st.empty()
    status_placeholder.info(st.session_state.status_message)
with status_col2:
    progress_placeholder = st.empty()

# Input section
with st.container(border=True):
    st.markdown("### 📋 Enter Usernames")
    usernames_text = st.text_area(
        "One username per line",
        height=120,
        label_visibility="collapsed",
        placeholder="User1\nUser2\nUser3\n..."
    )

# Buttons
col_btns = st.columns(5)
with col_btns[0]:
    start_clicked = st.button(
        "▶ Start Check", type="primary", use_container_width=True,
        disabled=st.session_state.scraping_active
    )
with col_btns[1]:
    stop_clicked = st.button(
        "⏹ Stop", use_container_width=True,
        disabled=not st.session_state.scraping_active
    )
with col_btns[2]:
    clear_clicked = st.button(
        "🗑 Clear", use_container_width=True,
        disabled=st.session_state.scraping_active
    )
with col_btns[3]:
    export_clicked = st.button(
        "📥 Export Excel", use_container_width=True,
        disabled=len(st.session_state.results) == 0 or st.session_state.scraping_active
    )

# ─── Handle button actions ─────────────────────────────────────────────
if start_clicked:
    usernames = [u.strip() for u in usernames_text.splitlines() if u.strip()]
    if not usernames:
        st.warning("Please enter at least one username.")
    else:
        thread = threading.Thread(target=run_batch, args=(usernames,), daemon=True)
        thread.start()
        st.rerun()

if stop_clicked:
    st.session_state.stop_flag = True
    st.session_state.status_message = "Stopping..."
    st.rerun()

if clear_clicked:
    st.session_state.results = []
    st.session_state.checked_count = 0
    st.session_state.total_to_check = 0
    st.session_state.status_message = "Cleared"
    st.rerun()

if export_clicked and st.session_state.results:
    excel_bytes = generate_excel(st.session_state.results)
    st.download_button(
        label="💾 Download Excel File",
        data=excel_bytes,
        file_name="AAA_Checker_Results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ─── Progress ───────────────────────────────────────────────────────────
if st.session_state.total_to_check > 0:
    pct = int(
        (st.session_state.checked_count / st.session_state.total_to_check) * 100
    )
    progress_placeholder.progress(
        pct / 100,
        text=f"{st.session_state.checked_count}/{st.session_state.total_to_check} ({pct}%)"
    )

# ─── Auto-refresh while scraping ────────────────────────────────────────
if st.session_state.scraping_active:
    status_placeholder.info(
        f"🔄 {st.session_state.status_message} "
        f"({st.session_state.checked_count}/{st.session_state.total_to_check})"
    )
    time.sleep(0.3)
    st.rerun()
else:
    if st.session_state.status_message == "Completed ✓":
        status_placeholder.success(st.session_state.status_message)
    elif "Stopped" in st.session_state.status_message:
        status_placeholder.warning(st.session_state.status_message)
    elif st.session_state.status_message == "Cleared":
        status_placeholder.info("Ready — cleared")
    else:
        status_placeholder.info(st.session_state.status_message)

# ─── Results table ──────────────────────────────────────────────────────
if st.session_state.results:
    st.markdown("### 📊 Results")
    df = pd.DataFrame(st.session_state.results)
    # Reorder columns to match the original
    df = df[columns] if all(c in df.columns for c in columns) else df

    st.dataframe(
        df,
        use_container_width=True,
        height=400,
        column_config={
            col: st.column_config.TextColumn(col, width="medium")
            for col in df.columns
        },
        hide_index=True,
    )

    st.caption(f"Total: {len(st.session_state.results)} users checked")
else:
    if not st.session_state.scraping_active:
        st.info("No results yet. Enter usernames and click **Start Check**.")
